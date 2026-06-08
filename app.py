import os
import re
import io
import email
import shutil
import zipfile
import imaplib
import unicodedata
import base64
import csv
import xml.etree.ElementTree as ET
from datetime import datetime
from email.header import decode_header

import streamlit as st
from openpyxl import load_workbook

# =====================================================
# CONFIGURACIÓN
# =====================================================
# RECOMENDADO PARA STREAMLIT CLOUD:
# En Secrets agrega:
# EMAIL_USER = "tu_correo@gmail.com"
# EMAIL_PASS = "tu_contraseña_de_aplicacion"
EMAIL_USER = st.secrets.get("EMAIL_USER", "difhermosillomunicipal@gmail.com")
EMAIL_PASS = st.secrets.get("EMAIL_PASS", "qcec ftus qhbw ckdi")

CARPETA_BASE = "datos_sistema"
CARPETA_FACTURAS = os.path.join(CARPETA_BASE, "facturas")
CARPETA_GENERADOS = os.path.join(CARPETA_BASE, "generados")
CARPETA_FORMATOS = "formatos"

TEMPLATE_REQUISICION = os.path.join(CARPETA_FORMATOS, "requisicion.xlsx")
TEMPLATE_COTIZACION = os.path.join(CARPETA_FORMATOS, "cotizacion.xlsx")
ARCHIVO_FOLIO = os.path.join(CARPETA_BASE, "folio.txt")
ARCHIVO_GASTOS = os.path.join(CARPETA_BASE, "movimientos_gastos.csv")

RFC_CALOTE = "AALK801205TH8"

os.makedirs(CARPETA_FACTURAS, exist_ok=True)
os.makedirs(CARPETA_GENERADOS, exist_ok=True)
os.makedirs(CARPETA_BASE, exist_ok=True)

AREAS_DIF = [
    "Administración",
    "Dirección General",
    "Procuraduría",
    "Adulto Mayor",
    "Discapacidad",
    "Jurídico",
    "Eventos",
    "Comunicación Social",
    "Voluntariado",
    "Almacén",
    "Otro"
]

# =====================================================
# HISTORIAL / REPORTES DE GASTO
# =====================================================
def inicializar_archivo_gastos():
    if not os.path.exists(ARCHIVO_GASTOS):
        with open(ARCHIVO_GASTOS, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "folio_sistema", "fecha_registro", "fecha_factura", "area",
                "responsable", "programa", "proveedor", "rfc_proveedor",
                "folio_factura", "uuid", "subtotal", "iva", "total", "concepto"
            ])


def guardar_movimiento_gasto(datos, folio, area, responsable, programa):
    inicializar_archivo_gastos()
    concepto_general = "; ".join([c.get("descripcion", "") for c in datos.get("conceptos", [])])[:900]

    with open(ARCHIVO_GASTOS, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            folio,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            datos.get("fecha", "")[:10],
            area,
            responsable,
            programa,
            datos.get("proveedor", ""),
            datos.get("rfc_proveedor", ""),
            datos.get("folio_factura", ""),
            datos.get("uuid", ""),
            datos.get("subtotal", 0),
            datos.get("iva", 0),
            datos.get("total", 0),
            concepto_general
        ])


def cargar_movimientos_gastos():
    inicializar_archivo_gastos()
    with open(ARCHIVO_GASTOS, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    for r in rows:
        for campo in ["total", "subtotal", "iva"]:
            try:
                r[campo] = float(r.get(campo, 0) or 0)
            except Exception:
                r[campo] = 0.0
    return rows


def uuids_ya_procesados():
    return {r.get("uuid", "") for r in cargar_movimientos_gastos() if r.get("uuid", "")}


def filtrar_por_fechas(rows, fecha_inicio, fecha_fin):
    filtrados = []
    for r in rows:
        fecha_txt = (r.get("fecha_factura") or r.get("fecha_registro") or "")[:10]
        try:
            fecha = datetime.strptime(fecha_txt, "%Y-%m-%d").date()
        except Exception:
            continue
        if fecha_inicio <= fecha <= fecha_fin:
            filtrados.append(r)
    return filtrados


def agrupar_total(rows, campo):
    resumen = {}
    for r in rows:
        llave = r.get(campo) or "Sin dato"
        resumen[llave] = resumen.get(llave, 0.0) + float(r.get("total", 0) or 0)
    return dict(sorted(resumen.items(), key=lambda x: x[1], reverse=True))

# =====================================================
# DISEÑO STREAMLIT
# =====================================================
st.set_page_config(
    page_title="Facturas CALOTE - DIF Hermosillo",
    page_icon="📄",
    layout="centered"
)

st.markdown("""
<style>
.stApp {
    background:
        radial-gradient(circle at top left, rgba(8,123,117,0.30), transparent 30%),
        radial-gradient(circle at bottom right, rgba(233,78,27,0.42), transparent 34%),
        linear-gradient(135deg, #EEF8F5 0%, #FFF7E7 50%, #F8C2A5 100%);
}
.block-container { padding-top: 25px; max-width: 980px; }
.header-title { color: #087B75; font-size: 38px; font-weight: 900; margin-top: 10px; text-align: center; }
.header-subtitle { color: #374151; font-size: 18px; margin-top: 6px; text-align: center; }
.info-card {
    background: linear-gradient(135deg, rgba(219,246,241,0.95), rgba(255,242,216,0.95));
    padding: 28px; border-radius: 20px; box-shadow: 0px 6px 18px rgba(0,0,0,0.12);
    margin-bottom: 26px; border-left: 8px solid #087B75; color: #123D3B; font-size: 17px;
}
.info-title { font-size: 22px; font-weight: 900; color: #087B75; margin-bottom: 12px; }
.result-card {
    background: linear-gradient(135deg, #E4F7F3, #FFF7EA);
    padding: 22px; border-radius: 18px; box-shadow: 0px 5px 15px rgba(0,0,0,0.10);
    margin-top: 18px; margin-bottom: 12px; border-left: 7px solid #E87522;
}
.pending-card {
    background: rgba(255,255,255,0.86);
    padding: 20px; border-radius: 18px; box-shadow: 0px 5px 15px rgba(0,0,0,0.10);
    margin-top: 18px; margin-bottom: 16px; border-left: 7px solid #087B75;
}
.stButton > button {
    background: linear-gradient(90deg, #E94E1B, #F2B233); color: white; border: none;
    border-radius: 16px; padding: 15px 22px; font-size: 18px; font-weight: 900;
    width: 70%; display: block; margin: 0 auto; box-shadow: 0px 6px 16px rgba(0,0,0,0.22);
}
.stButton > button:hover { background: linear-gradient(90deg, #D94316, #E6A222); transform: scale(1.01); }
.stDownloadButton > button {
    background: linear-gradient(90deg, #087B75, #14A39A); color: white; border: none;
    border-radius: 14px; padding: 13px 18px; font-size: 16px; font-weight: 800; width: 100%;
}
.footer { text-align: center; color: #087B75; font-size: 14px; margin-top: 35px; font-weight: 700; }
</style>
""", unsafe_allow_html=True)

# =====================================================
# FUNCIONES GENERALES
# =====================================================
def limpiar_nombre(nombre):
    if not nombre:
        return "sin_nombre"
    nombre = str(nombre)
    nombre = nombre.replace("Ñ", "N").replace("ñ", "n")
    nombre = unicodedata.normalize("NFKD", nombre)
    nombre = "".join(c for c in nombre if not unicodedata.combining(c))
    nombre = re.sub(r'[\\/*?:"<>|]', "_", nombre)
    nombre = nombre.encode("ascii", "ignore").decode("ascii")
    return nombre.strip() or "sin_nombre"


def money(valor):
    try:
        return float(valor)
    except Exception:
        return 0.0


def escribir(ws, celda, valor):
    for rango in ws.merged_cells.ranges:
        if celda in rango:
            ws.cell(rango.min_row, rango.min_col).value = valor
            return
    ws[celda] = valor


def limpiar_luis_carlos(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "LUIS CARLOS RUIZ" in cell.value.upper():
                cell.value = ""


def obtener_folio():
    if not os.path.exists(ARCHIVO_FOLIO):
        with open(ARCHIVO_FOLIO, "w", encoding="utf-8") as f:
            f.write("1721")

    with open(ARCHIVO_FOLIO, "r", encoding="utf-8") as f:
        folio = int(f.read().strip())

    with open(ARCHIVO_FOLIO, "w", encoding="utf-8") as f:
        f.write(str(folio + 1))

    return folio


def decodificar(texto):
    if not texto:
        return ""
    try:
        partes = decode_header(texto)
        resultado = ""
        for parte, codificacion in partes:
            if isinstance(parte, bytes):
                try:
                    resultado += parte.decode(codificacion or "utf-8", errors="ignore")
                except Exception:
                    resultado += parte.decode("latin-1", errors="ignore")
            else:
                resultado += str(parte)
        return resultado
    except Exception:
        return str(texto)

# =====================================================
# LEER XML CFDI
# =====================================================
def leer_xml_cfdi(ruta_xml):
    ns = {
        "cfdi": "http://www.sat.gob.mx/cfd/4",
        "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"
    }

    tree = ET.parse(ruta_xml)
    root = tree.getroot()

    emisor = root.find("cfdi:Emisor", ns)
    receptor = root.find("cfdi:Receptor", ns)
    conceptos = root.find("cfdi:Conceptos", ns)
    timbre = root.find(".//tfd:TimbreFiscalDigital", ns)

    datos = {
        "uuid": timbre.attrib.get("UUID", "") if timbre is not None else "",
        "serie": root.attrib.get("Serie", ""),
        "folio_factura": root.attrib.get("Folio", ""),
        "fecha": root.attrib.get("Fecha", ""),
        "subtotal": money(root.attrib.get("SubTotal", 0)),
        "total": money(root.attrib.get("Total", 0)),
        "moneda": root.attrib.get("Moneda", "MXN"),
        "proveedor": emisor.attrib.get("Nombre", "") if emisor is not None else "",
        "rfc_proveedor": emisor.attrib.get("Rfc", "") if emisor is not None else "",
        "receptor": receptor.attrib.get("Nombre", "") if receptor is not None else "",
        "rfc_receptor": receptor.attrib.get("Rfc", "") if receptor is not None else "",
        "conceptos": []
    }

    datos["iva"] = datos["total"] - datos["subtotal"]

    if conceptos is not None:
        for c in conceptos.findall("cfdi:Concepto", ns):
            datos["conceptos"].append({
                "cantidad": money(c.attrib.get("Cantidad", 0)),
                "unidad": c.attrib.get("Unidad", c.attrib.get("ClaveUnidad", "Pieza")),
                "codigo": c.attrib.get("NoIdentificacion", c.attrib.get("ClaveProdServ", "")),
                "descripcion": c.attrib.get("Descripcion", ""),
                "precio": money(c.attrib.get("ValorUnitario", 0)),
                "importe": money(c.attrib.get("Importe", 0))
            })

    return datos

# =====================================================
# REQUISICIÓN / COTIZACIÓN
# =====================================================
def llenar_requisicion(datos, folio, carpeta_salida, area="ADMINISTRATIVA", responsable="", programa="GENERALES"):
    wb = load_workbook(TEMPLATE_REQUISICION)
    ws = wb.active

    fecha = datetime.now()

    escribir(ws, "J6", folio)
    escribir(ws, "J10", fecha.day)
    escribir(ws, "K10", fecha.month)
    escribir(ws, "M10", fecha.year)

    escribir(ws, "D9", area.upper())
    escribir(ws, "D10", (programa or "GENERALES").upper())
    escribir(ws, "D11", (responsable or "INMEDIATA").upper())

    fila = 15
    for concepto in datos["conceptos"]:
        escribir(ws, f"B{fila}", concepto["cantidad"])
        escribir(ws, f"D{fila}", concepto["unidad"])
        escribir(ws, f"E{fila}", concepto["descripcion"])
        escribir(ws, f"H{fila}", f"FACTURA {datos['folio_factura']} / {datos['proveedor']}")
        fila += 1

    escribir(ws, "B51", f"COMPRA SEGÚN FACTURA {datos['folio_factura']} DE {datos['proveedor']} TOTAL ${datos['total']:,.2f}")
    limpiar_luis_carlos(ws)

    salida = os.path.join(carpeta_salida, f"REQUISICION_{folio}.xlsx")
    wb.save(salida)
    return salida


def llenar_cotizacion(datos, folio, carpeta_salida):
    wb = load_workbook(TEMPLATE_COTIZACION)
    ws = wb.active

    fecha = datetime.now().strftime("%d/%m/%Y")
    escribir(ws, "I1", f"Cotización #{folio} {fecha}")

    fila = 14
    total_articulos = 0

    for concepto in datos["conceptos"]:
        escribir(ws, f"A{fila}", concepto["cantidad"])
        escribir(ws, f"B{fila}", concepto["codigo"])
        escribir(ws, f"C{fila}", concepto["descripcion"])
        escribir(ws, f"F{fila}", concepto["precio"])
        escribir(ws, f"G{fila}", concepto["importe"])
        escribir(ws, f"I{fila}", concepto["importe"])
        total_articulos += concepto["cantidad"]
        fila += 1

    escribir(ws, "A35", f"Total de artículos:\n\n{int(total_articulos)}")
    escribir(ws, "I35", datos["subtotal"])
    escribir(ws, "I36", datos["iva"])
    escribir(ws, "I37", datos["total"])
    escribir(ws, "A37", f"CANTIDAD CON LETRA: TOTAL ${datos['total']:,.2f} M.N.")

    salida = os.path.join(carpeta_salida, f"COTIZACION_{folio}.xlsx")
    wb.save(salida)
    return salida

# =====================================================
# PASO 1: BUSCAR FACTURAS PENDIENTES SIN GENERAR DOCUMENTOS
# =====================================================
def buscar_facturas_pendientes():
    procesados = uuids_ya_procesados()

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(EMAIL_USER, EMAIL_PASS)
    mail.select("inbox")

    status, data = mail.search(None, 'TEXT "CALOTE"')
    if status != "OK":
        mail.logout()
        return []

    ids = data[0].split()
    pendientes = []
    vistos_en_busqueda = set()

    for num in ids:
        status, msg_data = mail.fetch(num, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(msg_data[0][1])
        asunto = limpiar_nombre(decodificar(msg.get("Subject")) or "correo")
        fecha_correo = datetime.now().strftime("%Y%m%d_%H%M%S")

        carpeta_correo = os.path.join(
            CARPETA_FACTURAS,
            limpiar_nombre(f"{fecha_correo}_{num.decode(errors='ignore')}_{asunto[:40]}")
        )
        os.makedirs(carpeta_correo, exist_ok=True)

        archivos_xml = []
        archivos_pdf = []

        for parte in msg.walk():
            if parte.get_content_disposition() == "attachment":
                nombre = decodificar(parte.get_filename())
                if not nombre:
                    continue

                nombre_limpio = limpiar_nombre(nombre)
                ruta = os.path.join(carpeta_correo, nombre_limpio)

                with open(ruta, "wb") as f:
                    f.write(parte.get_payload(decode=True))

                if nombre_limpio.lower().endswith(".xml"):
                    archivos_xml.append(ruta)
                elif nombre_limpio.lower().endswith(".pdf"):
                    archivos_pdf.append(ruta)

        for xml in archivos_xml:
            try:
                datos = leer_xml_cfdi(xml)

                if datos["rfc_proveedor"] != RFC_CALOTE:
                    continue

                uuid = datos.get("uuid", "")
                llave = uuid or f"{datos.get('folio_factura','')}_{datos.get('total','')}_{xml}"

                if uuid and uuid in procesados:
                    continue
                if llave in vistos_en_busqueda:
                    continue
                vistos_en_busqueda.add(llave)

                pendientes.append({
                    "datos": datos,
                    "xml": xml,
                    "pdfs": archivos_pdf,
                    "asunto": asunto
                })

            except Exception as e:
                pendientes.append({"error": str(e), "xml": xml})

    mail.logout()
    return pendientes

# =====================================================
# PASO 2: GENERAR DOCUMENTOS CON ÁREA POR FACTURA
# =====================================================
def generar_documentos_factura(factura, area, responsable, programa):
    datos = factura["datos"]
    folio = obtener_folio()

    carpeta_salida = os.path.join(
        CARPETA_GENERADOS,
        limpiar_nombre(f"FOLIO_{folio}_{datos['folio_factura']}")
    )
    os.makedirs(carpeta_salida, exist_ok=True)

    xml_copiado = os.path.join(carpeta_salida, os.path.basename(factura["xml"]))
    shutil.copy(factura["xml"], xml_copiado)

    pdf_copiado = None
    pdfs_copiados = []
    for pdf in factura.get("pdfs", []):
        destino_pdf = os.path.join(carpeta_salida, os.path.basename(pdf))
        shutil.copy(pdf, destino_pdf)
        pdfs_copiados.append(destino_pdf)
        if pdf_copiado is None:
            pdf_copiado = destino_pdf

    req = llenar_requisicion(datos, folio, carpeta_salida, area, responsable, programa)
    cot = llenar_cotizacion(datos, folio, carpeta_salida)
    guardar_movimiento_gasto(datos, folio, area, responsable, programa)

    return {
        "folio": folio,
        "factura": datos["folio_factura"],
        "proveedor": datos["proveedor"],
        "total": datos["total"],
        "area": area,
        "responsable": responsable,
        "programa": programa,
        "requisicion": req,
        "cotizacion": cot,
        "xml": xml_copiado,
        "pdf": pdf_copiado,
        "pdfs": pdfs_copiados
    }

# =====================================================
# CREAR ZIP
# =====================================================
def crear_zip_expediente(r):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(r["requisicion"], os.path.basename(r["requisicion"]))
        zipf.write(r["cotizacion"], os.path.basename(r["cotizacion"]))
        zipf.write(r["xml"], os.path.basename(r["xml"]))

        for pdf in r.get("pdfs", []):
            if pdf and os.path.exists(pdf):
                zipf.write(pdf, os.path.basename(pdf))
        if not r.get("pdfs") and r.get("pdf") and os.path.exists(r["pdf"]):
            zipf.write(r["pdf"], os.path.basename(r["pdf"]))

    zip_buffer.seek(0)
    return zip_buffer

# =====================================================
# PANTALLA PRINCIPAL
# =====================================================
if "facturas_pendientes" not in st.session_state:
    st.session_state["facturas_pendientes"] = []
if "resultados_generados" not in st.session_state:
    st.session_state["resultados_generados"] = []

if os.path.exists("logo_dif.png"):
    with open("logo_dif.png", "rb") as img_file:
        logo_base64 = base64.b64encode(img_file.read()).decode()

    st.markdown(f"""
    <div style="display:flex;justify-content:center;align-items:center;margin-top:5px;margin-bottom:35px;">
        <div style="background:linear-gradient(135deg, rgba(216,245,240,0.95), rgba(255,235,205,0.95));
            padding:35px 90px;border-radius:25px;box-shadow:0 10px 25px rgba(0,0,0,0.12);
            border:2px solid rgba(255,255,255,0.85);text-align:center;">
            <img src="data:image/png;base64,{logo_base64}" width="420" style="display:block;margin:auto;">
        </div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.warning("No se encontró logo_dif.png. Guarda el logo PNG en la misma carpeta que app.py")

st.markdown("""
<div class="header-title">Sistema de Facturas</div>
<div class="header-subtitle">DIF Hermosillo | Generación automática de requisiciones y cotizaciones</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="info-card" style="max-width:900px;margin:auto;margin-bottom:26px;">
    <div class="info-title">ℹ️ Nuevo flujo de trabajo</div>
    Primero busca las facturas pendientes. Después asigna el área, responsable y programa a cada factura individualmente.
    Finalmente genera los documentos y el historial de gasto queda correcto por área.
</div>
""", unsafe_allow_html=True)

st.markdown("## 📥 Facturas pendientes")
col_buscar, col_limpiar = st.columns(2)
with col_buscar:
    if st.button("🔍 Buscar facturas pendientes"):
        with st.spinner("Buscando correos de CALOTE y leyendo XML..."):
            try:
                st.session_state["facturas_pendientes"] = buscar_facturas_pendientes()
                st.session_state["resultados_generados"] = []
            except Exception as e:
                st.error(f"Error al buscar facturas: {e}")
                st.session_state["facturas_pendientes"] = []
with col_limpiar:
    if st.button("🧹 Limpiar búsqueda"):
        st.session_state["facturas_pendientes"] = []
        st.session_state["resultados_generados"] = []

pendientes = st.session_state.get("facturas_pendientes", [])

if pendientes:
    facturas_validas = [f for f in pendientes if "datos" in f]
    errores = [f for f in pendientes if "error" in f]

    if errores:
        for e in errores:
            st.error(f"Error XML: {e.get('xml', '')} - {e.get('error', '')}")

    if facturas_validas:
        st.success(f"Se encontraron {len(facturas_validas)} factura(s) pendiente(s). Asigna el área a cada una.")

        for i, f in enumerate(facturas_validas):
            datos = f["datos"]
            st.markdown(f"""
            <div class="pending-card">
                <h3 style="color:#087B75;margin-bottom:5px;">📄 Factura {datos.get('folio_factura','')}</h3>
                <p><b>Proveedor:</b> {datos.get('proveedor','')}</p>
                <p><b>Fecha factura:</b> {datos.get('fecha','')[:10]}</p>
                <p><b>Total:</b> ${datos.get('total',0):,.2f}</p>
                <p><b>UUID:</b> {datos.get('uuid','')}</p>
            </div>
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                st.selectbox("Área solicitante", AREAS_DIF, key=f"area_factura_{i}")
                st.text_input("Responsable o solicitante", key=f"responsable_factura_{i}", placeholder="Ejemplo: Coordinación Administrativa")
            with c2:
                st.text_input("Programa / Centro de costo", value="GENERALES", key=f"programa_factura_{i}")
                st.checkbox("Generar esta factura", value=True, key=f"generar_factura_{i}")

        if st.button("✅ Generar documentos seleccionados"):
            resultados = []
            with st.spinner("Generando requisiciones, cotizaciones y expediente ZIP..."):
                for i, f in enumerate(facturas_validas):
                    if not st.session_state.get(f"generar_factura_{i}", True):
                        continue
                    try:
                        area = st.session_state.get(f"area_factura_{i}", "Administración")
                        responsable = st.session_state.get(f"responsable_factura_{i}", "")
                        programa = st.session_state.get(f"programa_factura_{i}", "GENERALES")
                        r = generar_documentos_factura(f, area, responsable, programa)
                        resultados.append(r)
                    except Exception as e:
                        resultados.append({"error": str(e), "factura": f.get("datos", {}).get("folio_factura", "")})

            st.session_state["resultados_generados"] = resultados
            if resultados:
                st.success("Proceso terminado correctamente.")
            else:
                st.warning("No se seleccionó ninguna factura para generar.")
    else:
        st.warning("No se encontraron facturas válidas pendientes.")
else:
    st.info("Presiona 'Buscar facturas pendientes' para comenzar.")

resultados = st.session_state.get("resultados_generados", [])
if resultados:
    st.markdown("## 📦 Expedientes generados")
    for r in resultados:
        if "error" in r:
            st.error(f"Error al generar factura {r.get('factura','')}: {r.get('error','')}")
            continue

        st.markdown(f"""
        <div class="result-card">
            <h3 style="color:#087B75;">📄 Folio generado: {r['folio']}</h3>
            <p><b>Factura:</b> {r['factura']}</p>
            <p><b>Proveedor:</b> {r['proveedor']}</p>
            <p><b>Área:</b> {r['area']}</p>
            <p><b>Responsable:</b> {r['responsable']}</p>
            <p><b>Programa:</b> {r['programa']}</p>
            <p><b>Total:</b> ${r['total']:,.2f}</p>
        </div>
        """, unsafe_allow_html=True)

        zip_buffer = crear_zip_expediente(r)
        st.download_button(
            label="📦 Descargar expediente completo ZIP",
            data=zip_buffer,
            file_name=f"EXPEDIENTE_{r['folio']}.zip",
            mime="application/zip",
            key=f"zip_{r['folio']}"
        )

# =====================================================
# REPORTES
# =====================================================
st.markdown("---")
st.markdown("## 📊 Reportes de gasto por área")

movimientos = cargar_movimientos_gastos()
hoy = datetime.now().date()
col_f1, col_f2 = st.columns(2)
with col_f1:
    fecha_inicio = st.date_input("Fecha inicial", hoy.replace(day=1))
with col_f2:
    fecha_fin = st.date_input("Fecha final", hoy)

filtrados = filtrar_por_fechas(movimientos, fecha_inicio, fecha_fin)
total_periodo = sum(float(r.get("total", 0) or 0) for r in filtrados)

c1, c2, c3 = st.columns(3)
c1.metric("Total del periodo", f"${total_periodo:,.2f}")
c2.metric("Facturas", len(filtrados))
res_area = agrupar_total(filtrados, "area")
area_mayor = next(iter(res_area.keys())) if res_area else "Sin datos"
c3.metric("Área con mayor gasto", area_mayor)

if filtrados:
    st.subheader("Gasto por área")
    st.bar_chart(res_area)

    st.subheader("Gasto por proveedor")
    res_proveedor = agrupar_total(filtrados, "proveedor")
    st.bar_chart(res_proveedor)

    st.subheader("Detalle de movimientos")
    st.dataframe(filtrados, use_container_width=True)

    salida_csv = io.StringIO()
    campos = [
        "folio_sistema", "fecha_factura", "area", "responsable", "programa",
        "proveedor", "folio_factura", "subtotal", "iva", "total", "concepto"
    ]
    writer = csv.DictWriter(salida_csv, fieldnames=campos)
    writer.writeheader()
    for r in filtrados:
        writer.writerow({c: r.get(c, "") for c in campos})

    st.download_button(
        "⬇️ Descargar reporte CSV",
        data=salida_csv.getvalue().encode("utf-8-sig"),
        file_name=f"reporte_gastos_{fecha_inicio}_a_{fecha_fin}.csv",
        mime="text/csv"
    )
else:
    st.info("Todavía no hay movimientos registrados para ese rango de fechas.")

st.markdown("""
<div class="footer">Sistema interno DIF Hermosillo · Facturas CALOTE</div>
""", unsafe_allow_html=True)
